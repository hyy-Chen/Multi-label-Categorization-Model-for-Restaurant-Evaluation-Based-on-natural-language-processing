from flask import Flask, request, jsonify
import pickle
from sklearn.naive_bayes import MultinomialNB
import jieba
import pandas as pd
from sklearn.feature_extraction.text import CountVectorizer
import json
from openpyxl import Workbook, load_workbook

# 读取模型
with open('data/classifiers.pickle', 'rb') as f:
    classifiers = pickle.load(f)
# 读取vectorizer对象
with open('data/vectorizer.pickle', 'rb') as f:
    vectorizer = pickle.load(f)
#  读取分词表
with open('data/stopwords.txt', 'r', encoding='utf-8') as f:
    stopwords = f.read().splitlines()


app = Flask(__name__)
mp = {"味道好": 2, "味道差": 3, "态度好": 4, "态度差": 5, "卫生状态": 6, "不新鲜": 7, "其他": 8}
labels = ["味道好", "味道差", "态度好", "态度差", "卫生状态", "不新鲜", "其他"]

# 使用模型分析数据 将分析的结果放入data/text.xlsx文件内保存，并返回对应text的json
def information_analysis_func(text):
    copy_text = text
    text = ' '.join(jieba.cut(text))
    text = ' '.join([word for word in text.split() if word not in stopwords])
    new_X = vectorizer.transform([text])
    js = {"text": copy_text, "labels": []}
    data = [copy_text]
    flg = False
    for label, clf in classifiers.items():
        proba = clf.predict_proba(new_X)[0, 1]
        # 如果概率大于0.5,就放入json里面
        print(f'{label}的概率为{proba:.2f}')
        flag = proba >= 0.5
        if flag:
            js["labels"].append(label)
            flg = True
            data.append(1)
        else:
            data.append(0)
    if not flg:
        js["labels"].append("其他")
        data.append(1)
    else:
        data.append(0)
    # 打开现有的XLSX文件
    workbook = load_workbook('./data.xlsx')
    worksheet = workbook.active
    # 写入文件
    worksheet.append(data)
    # 保存文件
    workbook.save('./data.xlsx')
    return js

def get_information_func(label):
    flag = False
    if label == "全部":
        flag = True
    # 打开Excel文件
    workbook = load_workbook('data.xlsx')
    # 选择第一个工作表
    data = []
    worksheet = workbook.active
    # 获取行数和列数
    max_row = worksheet.max_row
    for i in range(2, max_row+1):
        if flag or worksheet.cell(row=i, column=mp[label]).value == 1:
            js = {"text": worksheet.cell(row=i, column=1).value, "labels": []}
            for key, value in mp.items():
                if worksheet.cell(row=i, column=value).value == 1:
                    js["labels"].append(key)
            data.append(js)
    return data



# 测试
# print(information_analysis("这家店味道很好，服务员态度也很好，服务周到，下次还来"))
@app.route('/test', methods=['POST'])
def test():
    return jsonify([{'text': '这家店味道很好，服务员态度也很好，服务周到，下次还来', 'labels': ['味道好', '态度好']}, {'text': '这家店味道很好，服务员态度也很好，服务周到，下次还来', 'labels': ['味道好', '态度好']}, {'text': '这家店味道很好，服务员态度也很好，服务周到，下次还来', 'labels': ['味道好', '态度好']}])

@app.route('/information_analysis', methods=['POST'])
def information_analysis():
    # 在这里编写信息分析的代码
    data = request.get_json()
    data = information_analysis_func(data["text"])
    print(data)
    return jsonify(data)

@app.route('/get_information', methods=['POST'])
def get_information():
    # 在这里编写获取信息的代码
    data = request.get_json()
    result = get_information_func(data["label"])
    return jsonify(result)

@app.route('/get_information', methods=['GET'])
def get_information_number():
    # 打开Excel文件
    workbook = load_workbook('data.xlsx')
    # 选择第一个工作表
    data = []
    worksheet = workbook.active
    # 获取行数和列数
    max_row = worksheet.max_row
    # 全部数量
    data.append(max_row-1)
    for i in labels:
        data.append(0)
    for i in range(2, max_row+1):
        for j in range(len(labels)):
            if worksheet.cell(row=i, column=mp[labels[j]]).value == 1:
                data[j+1] = data[j+1]+1
    return jsonify({"data": data})


if __name__ == '__main__':
    app.run(host="0.0.0.0", port=8989)
